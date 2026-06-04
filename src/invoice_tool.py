# -*- coding: utf-8 -*-
"""
发票归档
功能：发票PDF识别、付款截图管理、合同附件管理、按月筛选、导出Excel
"""

import sys
import os
import re
import json
import shutil
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QTableWidget, QTableWidgetItem,
    QFileDialog, QMessageBox, QHeaderView, QStatusBar, QFrame,
    QProgressBar, QAbstractItemView, QDialog,
    QComboBox, QMenu, QInputDialog, QSizePolicy
)
from PyQt5.QtCore import Qt, QTimer, QMimeData, QUrl, QEvent
from PyQt5.QtGui import QColor, QDesktopServices, QDragEnterEvent, QDropEvent, QFont

from dialogs import (ImageViewerDialog, ContractManagerDialog,
                     SettingsDialog, DeleteConfirmDialog)
from worker import ParseWorker
from filters import (record_matches_filter, get_available_years,
                     get_available_inv_types, get_available_sellers)
from models import Invoice
from repository import InvoiceRepository
from utils import safe_float
from services.invoice_service import InvoiceService
from ui.icons import get as get_icon
from ui.theme import (TABLE_QSS, PROGRESS_QSS, SUMMARY_FRAME_QSS,
                       ACCENT, RED, GREEN, TEXT, TEXT_SEC, TEXT_DIM,
                       BORDER_LIGHT, MONO_FONT, FS_SM, FS, FS_LG, FS_XL)
from version import APP_VERSION
from logger import setup_logging, shutdown_logging, getLogger
log = getLogger(__name__)


# 表格列定义
COLUMNS = ["发票类型", "购买方名称", "纳税人识别号",
           "销售方名称", "金额(元)", "征收率", "税额(元)", "价税合计(元)",
           "发票号码", "开票日期", "企业号", "付款截图", "合同", "备注"]
TABLE_ROW_HEIGHT = 36
FREEZE_COL_WIDTH = 96  # 冻结操作列宽度
COL_IDX = {c: i for i, c in enumerate(COLUMNS)}

# 支持的文件扩展名
IMG_EXTS      = {'.png', '.jpg', '.jpeg', '.bmp', '.gif', '.webp', '.tiff', '.tif'}
CONTRACT_EXTS = {'.pdf', '.docx', '.doc', '.xlsx', '.xls'}  # 合同支持格式


class InvoiceApp(QMainWindow):

    def __init__(self):
        super().__init__()
        self.records = []
        self._duplicate_invoices = []
        self.pending_company = ""
        log.info("InvoiceApp 初始化开始")
        
        # 配置文件路径：%APPDATA%\lan-invoice\config.json
        appdata = os.environ.get('APPDATA', os.path.expanduser('~'))
        self._config_file = os.path.join(appdata, 'lan-invoice', 'config.json')
        
        # 先读取配置文件获取数据目录（直接存储 _data_dir，不再嵌套 data 子目录）
        self._data_dir = self._load_config_dir()
        self._data_file      = os.path.join(self._data_dir, "invoices_data.json")
        self._attachment_dir = self._data_dir
        self._screenshot_dir = os.path.join(self._data_dir, "screenshots")
        self._contract_dir   = os.path.join(self._data_dir, "contracts")
        os.makedirs(self._data_dir,       exist_ok=True)
        os.makedirs(self._screenshot_dir, exist_ok=True)
        os.makedirs(self._contract_dir,   exist_ok=True)

        self._repo = InvoiceRepository(self._data_file)
        self._svc  = InvoiceService(self._repo, self._attachment_dir,
                                     os.path.join(self._data_dir, "invoices"))

        self._filter_year        = None
        self._filter_month       = None
        self._filter_inv_type    = None   # 发票类型筛选
        self._filter_seller      = None   # 销售方名称筛选
        self._filter_company     = ""     # 企业号搜索（模糊匹配）
        self._filter_buyer       = ""     # 购买方名称/税号搜索（模糊匹配）
        self._show_advanced_filter = False

        # 拖拽模式：'pdf'=导入发票, 'screenshot'=添加截图, 'contract'=添加合同
        # 通过键盘修饰键区分：Alt=截图, Shift=合同, 无修饰=PDF
        self._drag_mode = None
        self._shown_records = []   # 当前筛选后的记录缓存，_rebuild_table 时更新

        self._init_ui()
        self.setAcceptDrops(True)
        self._load_data()
        log.info("InvoiceApp 初始化完成 | 版本=%s | 数据目录=%s",
                 APP_VERSION, self._data_dir)
        QTimer.singleShot(500, self._check_desktop_shortcut)

    # ── 记录辅助方法 ────────────────────────────

    def _init_record_fields(self, data):
        """统一初始化记录字段默认值；红票金额转负数"""
        if isinstance(data, Invoice):
            data.ensure_defaults()
            return
        # 兼容旧 dict（逐步淘汰）
        for field in ("pdf_path", "invoice_type", "seller_name", "remark"):
            data.setdefault(field, "")
        data.setdefault("screenshots", [])
        data.setdefault("contracts", [])
        data.setdefault("is_red", False)
        if data.get("is_red"):
            for f in ("amount", "tax_amount", "total"):
                v = data.get(f, "")
                if v and not str(v).startswith('-'):
                    data[f] = '-' + str(v)

    def _find_record_index(self, inv_no: str):
        """在 self.records 中按发票号码查找索引，返回 int 或 None"""
        if not inv_no:
            return None
        for i, r in enumerate(self.records):
            if r.get("invoice_no") == inv_no:
                return i
        return None

    # ── UI 构建 ─────────────────────────────────
    def _init_ui(self):
        self.setWindowTitle(f"发票归档 v{APP_VERSION}")
        self._set_app_icon()
        self.resize(1480, 820)
        self.setMinimumSize(1000, 640)

        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(12, 12, 12, 8)
        main_layout.setSpacing(12)

        # ── 工具栏第一行 ──────────────────────────
        top_bar = QHBoxLayout()
        top_bar.setSpacing(8)

        self.btn_open = QPushButton(" 导入发票PDF")
        self.btn_open.setIcon(get_icon('folder'))
        self.btn_open.setFixedHeight(36)
        self.btn_open.setToolTip("选择一个或多个PDF发票文件（也可直接拖拽PDF到窗口）")
        self.btn_open.clicked.connect(self.open_files)

        self.btn_clear = QPushButton(" 清空列表")
        self.btn_clear.setIcon(get_icon('clear'))
        self.btn_clear.setFixedHeight(36)
        self.btn_clear.clicked.connect(self.clear_records)

        self.btn_settings = QPushButton(" 设置")
        self.btn_settings.setIcon(get_icon('settings'))
        self.btn_settings.setFixedHeight(36)
        self.btn_settings.setToolTip("数据目录设置 / 软件另存")
        self.btn_settings.clicked.connect(self._open_settings)

        self.btn_export = QPushButton(" 导出 Excel")
        self.btn_export.setIcon(get_icon('export'))
        self.btn_export.setFixedHeight(36)
        self.btn_export.setStyleSheet(
            f"background:{ACCENT}; color:white; font-weight:bold; border:none;")
        self.btn_export.setCursor(Qt.PointingHandCursor)
        self.btn_export.clicked.connect(self.export_excel)

        top_bar.addWidget(self.btn_open)
        top_bar.addWidget(self.btn_clear)
        top_bar.addWidget(self.btn_settings)
        top_bar.addStretch()

        lbl = QLabel("企业号")
        lbl.setFixedWidth(50)
        self.edit_company = QLineEdit()
        self.edit_company.setPlaceholderText("输入后新导入发票自动填入")
        self.edit_company.setMinimumWidth(120)
        self.edit_company.setFixedHeight(32)
        self.edit_company.textChanged.connect(self._on_company_changed)

        self.btn_apply = QPushButton("应用到已选行")
        self.btn_apply.setFixedHeight(32)
        self.btn_apply.clicked.connect(self.apply_company_to_selected)

        top_bar.addWidget(lbl)
        top_bar.addWidget(self.edit_company, 1)
        top_bar.addWidget(self.btn_apply)
        top_bar.addSpacing(12)
        top_bar.addWidget(self.btn_export)
        main_layout.addLayout(top_bar)

        # ── 工具栏第二行：多维筛选 ───────────────
        filter_bar = QHBoxLayout()
        filter_bar.setSpacing(12)

        lbl_filter = QLabel("筛选：")
        lbl_filter.setStyleSheet(f"font-size:{FS}; font-weight:bold; color:{TEXT};")

        # 年份
        lbl_y = QLabel("年份")
        lbl_y.setStyleSheet(f"font-size:{FS_SM}; color:{TEXT_SEC};")
        self.combo_year = QComboBox()
        self.combo_year.setFixedWidth(90)
        self.combo_year.setFixedHeight(30)
        self.combo_year.addItem("全部", None)

        # 月份
        lbl_m = QLabel("月份")
        lbl_m.setStyleSheet(f"font-size:{FS_SM}; color:{TEXT_SEC};")
        self.combo_month = QComboBox()
        self.combo_month.setFixedWidth(80)
        self.combo_month.setFixedHeight(30)
        self.combo_month.addItem("全部", None)
        for i in range(1, 13):
            self.combo_month.addItem(f"{i:02d} 月", i)

        # 高级筛选切换按钮
        self.btn_advanced_filter = QPushButton("▸ 高级筛选")
        self.btn_advanced_filter.setFixedHeight(30)
        self.btn_advanced_filter.setStyleSheet(f"font-size:{FS_SM}; color:{ACCENT}; border:none; background:transparent;")
        self.btn_advanced_filter.setCursor(Qt.PointingHandCursor)
        self.btn_advanced_filter.clicked.connect(self._toggle_advanced_filter)

        self.btn_filter = QPushButton("筛 选")
        self.btn_filter.setFixedHeight(30)
        self.btn_filter.setFixedWidth(70)
        self.btn_filter.clicked.connect(self._apply_filter)

        self.btn_reset = QPushButton("重置")
        self.btn_reset.setFixedHeight(30)
        self.btn_reset.setFixedWidth(60)
        self.btn_reset.clicked.connect(self._reset_filter)

        self.lbl_filter_hint = QLabel("")
        self.lbl_filter_hint.setStyleSheet(f"color:{ACCENT}; font-size:{FS_SM}; font-weight:bold;")

        filter_bar.addWidget(lbl_filter)
        filter_bar.addWidget(lbl_y)
        filter_bar.addWidget(self.combo_year)
        filter_bar.addWidget(lbl_m)
        filter_bar.addWidget(self.combo_month)
        filter_bar.addWidget(self.btn_advanced_filter)
        filter_bar.addStretch()
        filter_bar.addWidget(self.btn_filter)
        filter_bar.addWidget(self.btn_reset)
        filter_bar.addWidget(self.lbl_filter_hint)
        main_layout.addLayout(filter_bar)

        # ── 高级筛选面板（默认隐藏）───────────────
        self._advanced_filter_frame = QFrame()
        self._advanced_filter_frame.setVisible(False)
        adv_layout = QHBoxLayout(self._advanced_filter_frame)
        adv_layout.setContentsMargins(0, 0, 0, 0)
        adv_layout.setSpacing(10)

        # 发票类型
        lbl_type = QLabel("发票类型")
        lbl_type.setStyleSheet(f"font-size:{FS_SM}; color:{TEXT_SEC};")
        self.combo_inv_type = QComboBox()
        self.combo_inv_type.setMinimumWidth(100)
        self.combo_inv_type.setFixedHeight(30)
        self.combo_inv_type.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.combo_inv_type.addItem("全部", None)

        # 销售方名称
        lbl_seller = QLabel("销售方")
        lbl_seller.setStyleSheet(f"font-size:{FS_SM}; color:{TEXT_SEC};")
        self.combo_seller = QComboBox()
        self.combo_seller.setMinimumWidth(120)
        self.combo_seller.setFixedHeight(30)
        self.combo_seller.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.combo_seller.addItem("全部", None)

        # 购买方名称/税号搜索
        lbl_buyer_search = QLabel("购买方")
        lbl_buyer_search.setStyleSheet(f"font-size:{FS_SM}; color:{TEXT_SEC};")
        self.edit_buyer_search = QLineEdit()
        self.edit_buyer_search.setPlaceholderText("名称或税号")
        self.edit_buyer_search.setMinimumWidth(120)
        self.edit_buyer_search.setFixedHeight(30)
        self.edit_buyer_search.returnPressed.connect(self._apply_filter)

        # 企业号搜索
        lbl_company_search = QLabel("企业号")
        lbl_company_search.setStyleSheet(f"font-size:{FS_SM}; color:{TEXT_SEC};")
        self.edit_company_search = QLineEdit()
        self.edit_company_search.setPlaceholderText("输入企业号搜索")
        self.edit_company_search.setMinimumWidth(100)
        self.edit_company_search.setFixedHeight(30)
        self.edit_company_search.returnPressed.connect(self._apply_filter)

        adv_layout.addWidget(lbl_type)
        adv_layout.addWidget(self.combo_inv_type, 1)
        adv_layout.addWidget(lbl_seller)
        adv_layout.addWidget(self.combo_seller, 1)
        adv_layout.addWidget(lbl_buyer_search)
        adv_layout.addWidget(self.edit_buyer_search, 2)
        adv_layout.addWidget(lbl_company_search)
        adv_layout.addWidget(self.edit_company_search, 2)
        adv_layout.addStretch()
        main_layout.addWidget(self._advanced_filter_frame)

        # ── 进度条 ───────────────────────────────
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(6)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setStyleSheet(PROGRESS_QSS)
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)

        # ── 统计汇总栏 ───────────────────────────
        self.summary_frame = QFrame()
        self.summary_frame.setFrameShape(QFrame.StyledPanel)
        self.summary_frame.setStyleSheet(SUMMARY_FRAME_QSS)
        sum_layout = QHBoxLayout(self.summary_frame)
        sum_layout.setContentsMargins(20, 8, 20, 8)
        sum_layout.setSpacing(48)

        self.lbl_count     = self._stat_label("发票总数", "0 张")
        self.lbl_total_amt = self._stat_label("金额合计", "¥ 0.00")
        self.lbl_total_tax = self._stat_label("税额合计", "¥ 0.00")
        self.lbl_total_all = self._stat_label("价税合计", "¥ 0.00")

        for w in [self.lbl_count, self.lbl_total_amt, self.lbl_total_tax, self.lbl_total_all]:
            sum_layout.addWidget(w)
        sum_layout.addStretch()
        main_layout.addWidget(self.summary_frame)

        # ── 主表格 ───────────────────────────────
        self.table = QTableWidget()
        self.table.setColumnCount(len(COLUMNS))
        self.table.setHorizontalHeaderLabels(COLUMNS)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.verticalHeader().setDefaultSectionSize(TABLE_ROW_HEIGHT)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        header = self.table.horizontalHeader()
        # 固定列（像素宽度不变）
        fixed_cols = {4: 88, 5: 55, 6: 88, 7: 98, 11: 90, 12: 90}
        # 弹性列（最小宽度）
        stretch_cols = {0: 100, 1: 130, 2: 130, 3: 130,
                        8: 110, 9: 90, 10: 90, 13: 80}
        for col, width in fixed_cols.items():
            header.setSectionResizeMode(col, QHeaderView.Fixed)
            self.table.setColumnWidth(col, width)
        for col, width in stretch_cols.items():
            header.setSectionResizeMode(col, QHeaderView.Interactive)
            self.table.setColumnWidth(col, width)  # 初始宽度

        # 弹性列按初始宽度等比分配（Qt 5 无 setStretchFactor，用 resizeEvent 实现）
        self._stretch_cols = stretch_cols
        self._stretch_factors = {col: w // 10 for col, w in stretch_cols.items()}
        self._stretch_total_weight = sum(self._stretch_factors.values())
        self._stretch_fixed_total = sum(fixed_cols.values())

        self._resize_timer = QTimer(self.table)
        self._resize_timer.setSingleShot(True)
        self._resize_timer.timeout.connect(self._resize_stretch_cols)

        self._table_resize_orig = self.table.resizeEvent
        def _on_table_resize(event):
            self._table_resize_orig(event)
            self._resize_timer.start(80)
        self.table.resizeEvent = _on_table_resize

        self.table.setStyleSheet(TABLE_QSS)

        # ── 表格区域：主表格横向滚动，右侧操作列冻结 ──
        table_area = QWidget()
        table_area_layout = QHBoxLayout(table_area)
        table_area_layout.setContentsMargins(0, 0, 0, 0)
        table_area_layout.setSpacing(0)
        table_area_layout.addWidget(self.table, 1)

        self._freeze_table = QTableWidget(table_area)
        self._freeze_table.setColumnCount(1)
        self._freeze_table.setHorizontalHeaderLabels(["操作"])
        self._freeze_table.setFixedWidth(FREEZE_COL_WIDTH)
        self._freeze_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self._freeze_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._freeze_table.verticalHeader().hide()
        self._freeze_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self._freeze_table.setColumnWidth(0, FREEZE_COL_WIDTH)
        self._freeze_table.verticalHeader().setDefaultSectionSize(TABLE_ROW_HEIGHT)
        self._freeze_table.setShowGrid(False)
        self._freeze_table.setFocusPolicy(Qt.NoFocus)
        self._freeze_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._freeze_table.setSelectionMode(QAbstractItemView.NoSelection)
        self._freeze_table.setStyleSheet(
            TABLE_QSS +
            "QTableWidget { border: none; border-left: 1px solid " + BORDER_LIGHT + "; }"
            "QHeaderView::section { border-right: none; }"
        )
        self._freeze_table.hide()
        table_area_layout.addWidget(self._freeze_table, 0)
        main_layout.addWidget(table_area)

        # 同步垂直滚动
        self.table.verticalScrollBar().valueChanged.connect(
            self._freeze_table.verticalScrollBar().setValue)
        self._freeze_table.verticalScrollBar().valueChanged.connect(
            self.table.verticalScrollBar().setValue)

        # ── 空状态引导 ────────────────────────────
        self._empty_overlay = QLabel(self.table.viewport())
        self._empty_overlay.setAlignment(Qt.AlignCenter)
        self._empty_overlay.setWordWrap(True)
        self._empty_overlay.setStyleSheet(
            f"color:{TEXT_DIM}; font-size:13px; background:transparent;"
            "padding:40px;"
        )
        self._empty_overlay.setText(
            "拖拽 PDF 文件到此处导入发票\n"
            "或点击「打开」按钮选择文件"
        )
        self._empty_overlay.hide()
        # 跟随 viewport 居中
        self._recenter_empty_overlay()

        # ── 状态栏 ───────────────────────────────
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage(
            "就绪 — 拖拽 PDF 导入发票 | 选中行后拖拽图片添加截图 | 选中行后拖拽合同文件添加合同 | Ctrl+V 粘贴截图/合同")

        self._set_global_style()
        self._save_locked = False
        self.table.cellChanged.connect(self._on_cell_changed)
        self.table.clicked.connect(self._on_table_clicked)
        # 安装 viewport 事件过滤器：点击已选中行取消选中
        self.table.viewport().installEventFilter(self)

        # 全局按钮手型光标
        for btn in self.findChildren(QPushButton):
            btn.setCursor(Qt.PointingHandCursor)

    def _stat_label(self, title, value):
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(1)
        lbl_t = QLabel(title)
        lbl_t.setStyleSheet(f"color:{TEXT_SEC}; font-size:{FS_SM};")
        lbl_v = QLabel(value)
        lbl_v.setStyleSheet(f"color:{ACCENT}; font-size:17px; font-weight:bold;")
        v.addWidget(lbl_t)
        v.addWidget(lbl_v)
        w._value_label = lbl_v
        return w

    def _resize_stretch_cols(self):
        available = self.table.viewport().width()
        free = available - self._stretch_fixed_total
        if free <= 0:
            return
        # 最大余数法：避免取整误差累积导致右侧空白
        raw = {}
        remainders = []
        for col, min_w in self._stretch_cols.items():
            exact = free * self._stretch_factors[col] / self._stretch_total_weight
            raw[col] = max(min_w, int(exact))
            remainders.append((col, exact - int(exact)))
        remainders.sort(key=lambda x: x[1], reverse=True)
        for i in range(free - sum(raw.values())):
            raw[remainders[i][0]] += 1
        for col, w in raw.items():
            self.table.setColumnWidth(col, w)

    def _set_global_style(self):
        from ui.theme import GLOBAL_QSS
        self.setStyleSheet(GLOBAL_QSS)

    def _set_app_icon(self):
        from PyQt5.QtGui import QIcon
        import os
        # 打包后从资源路径加载，开发时从文件加载
        paths = [
            os.path.join(os.path.dirname(__file__), "ui", "icons", "app_icon.png"),
            os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), "ui", "icons", "app_icon.png"),
            "icon.ico",
        ]
        for p in paths:
            if os.path.exists(p):
                self.setWindowIcon(QIcon(p))
                return

    # ── 筛选条件 ─────────────────────────────────
    def _get_available_years(self):
        return get_available_years(self.records)

    def _get_available_inv_types(self):
        return get_available_inv_types(self.records)

    def _get_available_sellers(self):
        return get_available_sellers(self.records)

    def _refresh_year_combo(self):
        current = self.combo_year.currentData()
        self.combo_year.blockSignals(True)
        self.combo_year.clear()
        self.combo_year.addItem("全部", None)
        for y in self._get_available_years():
            self.combo_year.addItem(str(y), y)
        idx = self.combo_year.findData(current)
        if idx >= 0:
            self.combo_year.setCurrentIndex(idx)
        self.combo_year.blockSignals(False)

    def _refresh_filter_combos(self):
        """动态刷新发票类型、销售方下拉选项（保留当前选中值）"""
        # 发票类型
        cur_type = self.combo_inv_type.currentData()
        self.combo_inv_type.blockSignals(True)
        self.combo_inv_type.clear()
        self.combo_inv_type.addItem("全部", None)
        for t in self._get_available_inv_types():
            self.combo_inv_type.addItem(t, t)
        idx = self.combo_inv_type.findData(cur_type)
        self.combo_inv_type.setCurrentIndex(idx if idx >= 0 else 0)
        self.combo_inv_type.blockSignals(False)

        # 销售方名称
        cur_seller = self.combo_seller.currentData()
        self.combo_seller.blockSignals(True)
        self.combo_seller.clear()
        self.combo_seller.addItem("全部", None)
        for s in self._get_available_sellers():
            self.combo_seller.addItem(s, s)
        idx = self.combo_seller.findData(cur_seller)
        self.combo_seller.setCurrentIndex(idx if idx >= 0 else 0)
        self.combo_seller.blockSignals(False)

        self._refresh_year_combo()

    def _apply_filter(self):
        self._filter_year     = self.combo_year.currentData()
        self._filter_month    = self.combo_month.currentData()
        self._filter_inv_type = self.combo_inv_type.currentData()
        self._filter_seller   = self.combo_seller.currentData()
        self._filter_buyer    = self.edit_buyer_search.text().strip()
        self._filter_company  = self.edit_company_search.text().strip()
        self._rebuild_table()
        parts = []
        if self._filter_year:
            parts.append(f"{self._filter_year}年")
        if self._filter_month:
            parts.append(f"{self._filter_month:02d}月")
        if self._filter_inv_type:
            parts.append(self._filter_inv_type)
        if self._filter_seller:
            parts.append(f"销售方:{self._filter_seller}")
        if self._filter_buyer:
            parts.append(f"购买方:{self._filter_buyer}")
        if self._filter_company:
            parts.append(f"企业号:{self._filter_company}")
        self.lbl_filter_hint.setText(f"当前筛选：{'  '.join(parts)}" if parts else "")

    def _toggle_advanced_filter(self):
        self._show_advanced_filter = not self._show_advanced_filter
        self._advanced_filter_frame.setVisible(self._show_advanced_filter)
        if self._show_advanced_filter:
            self.btn_advanced_filter.setText("▾ 高级筛选")
        else:
            self.btn_advanced_filter.setText("▸ 高级筛选")

    def _reset_filter(self):
        self._filter_year     = None
        self._filter_month    = None
        self._filter_inv_type = None
        self._filter_seller   = None
        self._filter_buyer    = ""
        self._filter_company  = ""
        self.combo_year.setCurrentIndex(0)
        self.combo_month.setCurrentIndex(0)
        self.combo_inv_type.setCurrentIndex(0)
        self.combo_seller.setCurrentIndex(0)
        self.edit_buyer_search.clear()
        self.edit_company_search.clear()
        self.lbl_filter_hint.setText("")
        self._rebuild_table()

    def _record_matches_filter(self, rec) -> bool:
        return record_matches_filter(
            rec, self._filter_year, self._filter_month,
            self._filter_inv_type, self._filter_seller,
            self._filter_buyer, self._filter_company)

    def _recenter_empty_overlay(self):
        """让空状态提示在 viewport 中居中"""
        if not hasattr(self, '_empty_overlay'):
            return
        vp = self.table.viewport()
        w = vp.width() - 60
        self._empty_overlay.setGeometry(30, (vp.height() - 100) // 2, w, 100)

    def _update_empty_state(self):
        """根据记录数显示/隐藏空状态引导"""
        if not hasattr(self, '_empty_overlay'):
            return
        if len(self.records) == 0:
            self._empty_overlay.show()
            self._recenter_empty_overlay()
        else:
            self._empty_overlay.hide()

    def _rebuild_table(self):
        self._save_locked = True
        scroll_val = self.table.verticalScrollBar().value()
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(0)
        self._shown_records = [r for r in self.records if self._record_matches_filter(r)]
        shown = self._shown_records
        for data in shown:
            self._insert_row(data, scroll=False)
        self.table.setUpdatesEnabled(True)
        self.table.verticalScrollBar().setValue(min(scroll_val, self.table.verticalScrollBar().maximum()))
        self._refresh_summary_from_list(shown)
        self._save_locked = False
        active = any([self._filter_year, self._filter_month,
                      self._filter_inv_type, self._filter_seller])
        if active:
            self.status.showMessage(f"筛选结果：显示 {len(shown)} 张 / 共 {len(self.records)} 张")
        elif len(self.records) == 0:
            self.status.showMessage(
                "开始使用：拖拽 PDF 发票到窗口即可自动识别归档 | "
                "Ctrl+O 导入 PDF | Ctrl+E 导出 Excel | Delete 删除选中行")
        self._update_empty_state()
        self._rebuild_freeze_table()

    # ── 数据持久化 ──────────────────────────────
    def _save_data(self):
        try:
            self._sync_records_from_table()
            self._repo.save(self.records)
            log.debug("数据已保存: %d 条", len(self.records))
        except (OSError, IOError) as e:
            log.error("数据保存失败: %s", e)
            self.status.showMessage(f"数据保存失败: {e}")

    def _sync_records_from_table(self):
        shown = self._shown_records
        for i in range(self.table.rowCount()):
            inv_no_item = self.table.item(i, COL_IDX["发票号码"])
            inv_no = inv_no_item.text() if inv_no_item else ""

            idx = self._find_record_index(inv_no)
            rec = self.records[idx] if idx is not None else None
            if rec is None and 0 <= i < len(shown):
                rec = shown[i]
            if rec is None:
                continue

            co_item = self.table.item(i, COL_IDX["企业号"])
            bk_item = self.table.item(i, COL_IDX["备注"])
            if co_item:
                rec["company"] = co_item.text()
            if bk_item and bk_item.text() != "✓":
                rec["remark"] = bk_item.text()

    def _load_config_dir(self):
        default_dir = os.path.join(os.path.dirname(self._config_file), "data")
        # 旧版配置迁移：项目根目录 config.json → %APPDATA%\lan-invoice\
        old_config = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "config.json"
        )
        if os.path.exists(old_config) and not os.path.exists(self._config_file):
            try:
                os.makedirs(os.path.dirname(self._config_file), exist_ok=True)
                shutil.copy2(old_config, self._config_file)
            except OSError:
                pass
        if not os.path.exists(self._config_file):
            return default_dir
        try:
            with open(self._config_file, "r", encoding="utf-8") as f:
                config = json.load(f)
                data_dir = config.get("data_dir", "")
                if data_dir and os.path.isdir(data_dir):
                    return data_dir
        except (OSError, json.JSONDecodeError):
            pass
        return default_dir

    def _save_config_dir(self, data_dir):
        try:
            os.makedirs(os.path.dirname(self._config_file), exist_ok=True)
            with open(self._config_file, "w", encoding="utf-8") as f:
                json.dump({"data_dir": data_dir}, f, ensure_ascii=False, indent=2)
        except OSError:
            pass

    def _load_data(self):
        invoices = self._repo.load()
        if not invoices:
            log.info("无历史数据，初始化空列表")
            return
        log.info("加载 %d 条历史记录", len(invoices))
        self.records = invoices
        self._save_locked = True
        self.table.setUpdatesEnabled(False)
        for inv in self.records:
            self._init_record_fields(inv)
            self._insert_row(inv, scroll=False)
        self.table.setUpdatesEnabled(True)
        self._shown_records = list(self.records)
        self._refresh_summary()
        self._refresh_filter_combos()
        self._save_locked = False
        self._update_empty_state()
        self._rebuild_freeze_table()
        self.status.showMessage(f"已自动加载 {len(self.records)} 条历史记录")

    # ── 键盘快捷键 ──────────────────────────────
    def keyPressEvent(self, event):
        """全局键盘快捷键"""
        mod = event.modifiers()
        key = event.key()

        if mod == Qt.ControlModifier:
            if key == Qt.Key_O:
                self.open_files()
                return
            elif key == Qt.Key_E:
                self.export_excel()
                return
            elif key == Qt.Key_F:
                # 聚焦第一个筛选控件
                self.edit_buyer_search.setFocus()
                self.edit_buyer_search.selectAll()
                return
            elif key == Qt.Key_V:
                # Ctrl+V 粘贴截图或合同
                rows = sorted(set(item.row() for item in self.table.selectedItems()))
                if not rows:
                    self.status.showMessage("请先选中一行，再按 Ctrl+V 粘贴截图或合同")
                    return
                self._paste_from_clipboard(rows[0])
                return

        if key == Qt.Key_Delete:
            # 删除选中行（仅在表格有焦点时）
            if self.table.hasFocus():
                self._delete_selected_rows()
                return
        elif key == Qt.Key_Escape:
            self._reset_filter()
            return

        super().keyPressEvent(event)

    # ── 拖拽支持 ────────────────────────────────
    def dragEnterEvent(self, e: QDragEnterEvent):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()
            self._show_drop_overlay(e.mimeData().urls())

    def dragLeaveEvent(self, e):
        self._hide_drop_overlay()

    def dropEvent(self, e: QDropEvent):
        self._hide_drop_overlay()
        urls = e.mimeData().urls()
        pdf_files  = []
        img_files  = []
        doc_files  = []

        for u in urls:
            path = u.toLocalFile()
            ext  = os.path.splitext(path)[1].lower()
            if ext == '.pdf':
                pdf_files.append(path)   # PDF 始终作为发票导入
            elif ext in IMG_EXTS:
                img_files.append(path)
            elif ext in CONTRACT_EXTS:
                doc_files.append(path)

        # PDF → 发票导入（不依赖选中状态）
        if pdf_files:
            self._start_parse(pdf_files)

        # 图片/文档 → 附件（需要选中行）
        rows = sorted(set(item.row() for item in self.table.selectedItems()))
        other_files = img_files + doc_files
        if other_files:
            if not rows:
                QMessageBox.information(
                    self, "提示",
                    "请先选中一行，再将图片/文档拖入作为附件添加。\n"
                    "PDF 文件将始终作为发票导入。"
                )
            else:
                self._add_attachments_from_paths(rows[0], other_files)

    def _show_drop_overlay(self, urls):
        if not hasattr(self, '_drop_overlay'):
            self._drop_overlay = QLabel(self)
            self._drop_overlay.setAlignment(Qt.AlignCenter)
            self._drop_overlay.setStyleSheet(
                "background: rgba(30, 111, 191, 200); color: white; "
                "font-size: 18px; font-weight: bold; border-radius: 12px;"
            )
        pdf_count = sum(1 for u in urls if u.toLocalFile().lower().endswith('.pdf'))
        other_count = len(urls) - pdf_count
        parts = []
        if pdf_count:
            parts.append(f"导入 {pdf_count} 个发票 PDF")
        if other_count:
            rows = set(item.row() for item in self.table.selectedItems())
            if rows:
                parts.append(f"添加 {other_count} 个附件到选中行")
            else:
                parts.append(f"⚠ 需选中行以添加 {other_count} 个附件")
        self._drop_overlay.setText("\n".join(parts))
        self._drop_overlay.setGeometry(0, 0, self.width(), self.height())
        self._drop_overlay.show()

    def _hide_drop_overlay(self):
        if hasattr(self, '_drop_overlay'):
            self._drop_overlay.hide()

    # ── 槽函数 ───────────────────────────────────
    def _on_company_changed(self, text):
        self.pending_company = text.strip()

    def _on_table_clicked(self, index):
        """点击表格行时，在状态栏显示当前行摘要信息"""
        row = index.row()
        try:
            rec = self._get_record_by_row(row)
            if rec:
                seller = rec.get("seller_name", "") or "—"
                date   = rec.get("invoice_date", "") or "—"
                total  = rec.get("total", "") or "—"
                self.status.showMessage(
                    f"第 {row + 1} 行 | {seller} | {date} | 价税合计：¥{total}"
                    "  ·  Ctrl+V 粘贴截图/合同"
                )
        except Exception:
            pass

    def _on_cell_changed(self, row, col):
        if self._save_locked:
            return
        header = self.table.horizontalHeaderItem(col).text()
        if header in ("企业号", "备注"):
            self._save_data()

    def closeEvent(self, event):
        log.info("应用关闭…")
        if hasattr(self, '_worker') and self._worker.isRunning():
            log.debug("等待后台线程结束…")
            self._worker.abort()
            if not self._worker.wait(3000):
                log.warning("后台线程未在3秒内结束")
        self._save_data()
        event.accept()

    def _open_settings(self):
        """打开设置对话框"""
        dlg = SettingsDialog(self, parent=self)
        dlg.exec_()

    def _check_desktop_shortcut(self):
        """首次启动检测桌面快捷方式，不存在则提示创建"""
        import subprocess
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        lnk_path = os.path.join(desktop, "发票归档.lnk")
        if os.path.exists(lnk_path):
            return
        reply = QMessageBox.question(
            self, "创建快捷方式",
            "是否在桌面创建「发票归档」快捷方式？\n\n"
            "创建后可从桌面直接双击启动软件。",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes
        )
        if reply != QMessageBox.Yes:
            return
        try:
            exe_path = sys.executable
            # 打包后 EXE 路径，开发时用脚本路径
            if not exe_path.endswith(".exe") or "python" in exe_path.lower():
                script = os.path.abspath(sys.argv[0]) if sys.argv[0] else __file__
                ps_cmd = (
                    f"$ws = New-Object -ComObject WScript.Shell; "
                    f"$sc = $ws.CreateShortcut('{lnk_path}'); "
                    f"$sc.TargetPath = 'pythonw'; "
                    f"$sc.Arguments = '{script}'; "
                    f"$sc.WorkingDirectory = '{os.path.dirname(script)}'; "
                    f"$sc.Save()"
                )
            else:
                ps_cmd = (
                    f"$ws = New-Object -ComObject WScript.Shell; "
                    f"$sc = $ws.CreateShortcut('{lnk_path}'); "
                    f"$sc.TargetPath = '{exe_path}'; "
                    f"$sc.Save()"
                )
            subprocess.run(
                ["powershell", "-NoProfile", "-Command", ps_cmd],
                capture_output=True, timeout=10
            )
            if os.path.exists(lnk_path):
                self.status.showMessage("桌面快捷方式已创建")
        except Exception:
            pass    # 静默失败，不影响主流程

    def open_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择发票PDF文件", "",
            "PDF文件 (*.pdf);;所有文件 (*)"
        )
        if files:
            self._start_parse(files)

    def clear_records(self):
        if not self.records:
            return
        reply = QMessageBox.question(self, "确认清空", "确定要清空所有记录吗？",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.records.clear()
            self.table.setRowCount(0)
            self._refresh_summary()
            self._refresh_filter_combos()
            self._save_data()
            self.status.showMessage("已清空")

    def apply_company_to_selected(self):
        rows = set(item.row() for item in self.table.selectedItems())
        if not rows:
            QMessageBox.information(self, "提示", "请先在表格中选中需要修改的行")
            return
        company = self.pending_company
        if not company:
            company, ok = QInputDialog.getText(self, "输入企业号", "企业号：")
            if not ok or not company:
                return
        col = COL_IDX["企业号"]
        shown = self._shown_records
        for row in rows:
            self.table.setItem(row, col, QTableWidgetItem(company))
            inv_no_item = self.table.item(row, COL_IDX["发票号码"])
            inv_no = inv_no_item.text() if inv_no_item else ""
            idx = self._find_record_index(inv_no)
            rec = self.records[idx] if idx is not None else None
            if rec is None and 0 <= row < len(shown):
                rec = shown[row]
            if rec:
                rec["company"] = company
        self.status.showMessage(f"已将企业号「{company}」应用到 {len(rows)} 行")
        self._save_data()

    # ── 解析流程 ─────────────────────────────────
    def _start_parse(self, files):
        self.btn_open.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status.showMessage(f"正在解析 {len(files)} 个文件...")
        self._parse_errors = []
        self._duplicate_invoices = []
        self._batch_count_before = len(self.records)

        # data_dir 传给 Worker，让文件复制在后台完成
        self._worker = ParseWorker(files, data_dir=self._data_dir)
        self._worker.progress.connect(self.progress_bar.setValue)
        self._worker.result_ready.connect(self._add_record_batch)
        self._worker.error_occurred.connect(self._on_parse_error)
        self._worker.finished.connect(self._parse_done)
        self._worker.start()

    def _on_parse_error(self, error_msg):
        self._parse_errors.append(error_msg)
        log.warning("解析错误: %s", error_msg)
        self.status.showMessage(f"解析错误: {error_msg}")

    # ── 批量导入专用槽（纯内存操作，不碰 UI）────────────────
    def _add_record_batch(self, data: dict):
        """后台每解析完一条调此槽；重复检测 + 写入"""
        inv = Invoice.from_dict(data)
        # 重复发票号检测：收集重复记录供结果摘要展示
        if inv.invoice_no and self._find_record_index(inv.invoice_no) is not None:
            self._duplicate_invoices.append(inv)
            return
        self._init_record_fields(inv)
        self.records.append(inv)

    def _insert_row(self, data: dict, scroll: bool = True):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setRowHeight(row, TABLE_ROW_HEIGHT)

        is_red = data.get("is_red", False)

        def cell(text, editable=False, fg=None, bg=None):
            it = QTableWidgetItem(str(text) if text else "")
            if not editable:
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)
            if fg:
                it.setForeground(QColor(fg))
            if bg:
                it.setBackground(QColor(bg))
            return it

        # 红票整行浅红背景
        row_bg = "#FFE4E4" if is_red else None


        # 发票类型：红票显示"● 红票-类型"，蓝票显示"● 类型"
        inv_type = data.get("invoice_type", "")
        if is_red:
            type_text = f"● 红票{'-' + inv_type if inv_type else ''}"
            type_fg   = RED
        else:
            type_text = f"● {inv_type}" if inv_type else ""
            type_fg   = ACCENT
        type_item = cell(type_text, fg=type_fg, bg=row_bg)
        self.table.setItem(row, COL_IDX["发票类型"], type_item)

        # 金额列：负数标红（红票金额已在入库时转负）
        def amount_cell(field):
            v = data.get(field, "")
            v = str(v) if v is not None else ""
            neg = v.startswith('-')
            return cell(v, fg=RED if neg else None, bg=row_bg if row_bg else ("#FFF0F0" if neg else None))

        self.table.setItem(row, COL_IDX["购买方名称"],   cell(data.get("buyer_name", ""), bg=row_bg))
        self.table.setItem(row, COL_IDX["纳税人识别号"],  cell(data.get("buyer_tax_id", ""), bg=row_bg))
        self.table.setItem(row, COL_IDX["销售方名称"],   cell(data.get("seller_name", ""), bg=row_bg))
        self.table.setItem(row, COL_IDX["金额(元)"],     amount_cell("amount"))
        self.table.setItem(row, COL_IDX["征收率"],       cell(data.get("tax_rate", ""), bg=row_bg))
        self.table.setItem(row, COL_IDX["税额(元)"],     amount_cell("tax_amount"))
        self.table.setItem(row, COL_IDX["价税合计(元)"],  amount_cell("total"))

        # 数字列使用等宽字体，保证金额对齐扫描
        _mono = QFont(MONO_FONT, 12)
        for c in (COL_IDX["金额(元)"], COL_IDX["征收率"], COL_IDX["税额(元)"], COL_IDX["价税合计(元)"]):
            it = self.table.item(row, c)
            if it:
                it.setFont(_mono)

        self.table.setItem(row, COL_IDX["发票号码"],      cell(data.get("invoice_no", ""), bg=row_bg))
        self.table.setItem(row, COL_IDX["开票日期"],      cell(data.get("invoice_date", ""), bg=row_bg))
        self.table.setItem(row, COL_IDX["企业号"],        cell(data.get("company", ""), editable=True, bg=row_bg))

        self._set_screenshot_cell(row, data)
        self._set_contract_cell(row, data)

        remark_val  = data.get("remark", "") or data.get("error", "") or "✓"
        remark_item = cell(remark_val, editable=True,
                           fg=RED if data.get("error") else (RED if is_red else (GREEN if remark_val == "✓" else TEXT)),
                           bg=row_bg)
        self.table.setItem(row, COL_IDX["备注"], remark_item)

        if scroll:
            self.table.scrollToBottom()

    # ── 冻结操作列 ────────────────────────────────

    def _rebuild_freeze_table(self):
        """重建冻结表格（与 _rebuild_table 同步调用）"""
        freeze = self._freeze_table
        freeze.setRowCount(0)
        shown = self._shown_records
        if not shown:
            freeze.hide()
            return
        freeze.setRowCount(len(shown))
        freeze.horizontalHeader().setFixedHeight(self.table.horizontalHeader().height())
        btn_style = "font-size:12px; padding:2px 4px; border:none; background:transparent;"

        for i, rec in enumerate(shown):
            freeze.setRowHeight(i, self.table.rowHeight(i))
            w = QWidget()
            lay = QHBoxLayout(w)
            lay.setContentsMargins(2, 2, 2, 2)
            lay.setSpacing(2)

            pdf_path = rec.get("pdf_path", "")
            if pdf_path and os.path.exists(pdf_path):
                btn_pdf = QPushButton("查看")
                btn_pdf.setFixedHeight(22)
                btn_pdf.setStyleSheet(btn_style + f"color:{ACCENT};")
                btn_pdf.setCursor(Qt.PointingHandCursor)
                btn_pdf.clicked.connect(self._make_pdf_handler(rec))
                lay.addWidget(btn_pdf)

            btn_del = QPushButton("删除")
            btn_del.setFixedHeight(22)
            btn_del.setStyleSheet(btn_style + f"color:{TEXT_DIM};")
            btn_del.setCursor(Qt.PointingHandCursor)
            btn_del.clicked.connect(self._make_delete_handler(rec))
            lay.addWidget(btn_del)
            lay.addStretch()
            freeze.setCellWidget(i, 0, w)

        freeze.show()
        freeze.verticalScrollBar().setValue(self.table.verticalScrollBar().value())

    def _make_pdf_handler(self, rec):
        return lambda: self._view_invoice_pdf_by_rec(rec)

    def _make_delete_handler(self, rec):
        return lambda: self._delete_record(rec)

    def _view_invoice_pdf_by_rec(self, rec):
        path = rec.get("pdf_path", "")
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "文件不存在", "PDF 文件不存在或已被移动。")
            return
        from ui.dialogs.pdf_viewer import PdfViewerDialog
        PdfViewerDialog(path, parent=self).exec_()

    def _delete_record(self, rec):
        if rec not in self.records:
            return
        for i in range(self.table.rowCount()):
            inv_no_item = self.table.item(i, COL_IDX["发票号码"])
            inv_no = inv_no_item.text() if inv_no_item else ""
            if inv_no == rec.get("invoice_no", ""):
                self.table.selectRow(i)
                break
        self._delete_selected_rows()

    def _view_invoice_pdf(self, row):
        """打开 PDF 预览窗口（内嵌下载/系统打开等操作）"""
        rec = self._get_record_by_row(row)
        if rec is None:
            return
        path = rec.get("pdf_path", "")
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "文件不存在", "PDF 文件不存在或已被移动。")
            return
        from ui.dialogs.pdf_viewer import PdfViewerDialog
        dlg = PdfViewerDialog(path, parent=self)
        dlg.exec_()

    # ── 截图单元格 ───────────────────────────────
    def _set_screenshot_cell(self, row, data):
        screenshots = data.get("screenshots", [])
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(4, 4, 4, 4)
        lay.setSpacing(4)

        if screenshots:
            lbl = QLabel(f"[{len(screenshots)}]")
            lbl.setStyleSheet(f"color:{ACCENT}; font-size:12px; font-weight:bold;")
            btn_v = QPushButton("查看")
            btn_v.setFixedHeight(24)
            btn_v.setFixedWidth(40)
            btn_v.setStyleSheet(f"font-size:11px; padding:1px 4px; color:{ACCENT}; border:none; background:transparent;")
            btn_v.clicked.connect(lambda _, r=row: self._view_screenshots(r))
            lay.addWidget(lbl)
            lay.addWidget(btn_v)
        else:
            lbl = QLabel("—")
            lbl.setStyleSheet(f"color:{TEXT_DIM}; font-size:12px;")
            lay.addWidget(lbl)

        btn_add = QPushButton("＋")
        btn_add.setFixedHeight(24)
        btn_add.setFixedWidth(26)
        btn_add.setToolTip("添加付款截图")
        btn_add.setStyleSheet(f"font-size:13px; padding:0; color:{ACCENT}; border:none; background:transparent;")
        btn_add.clicked.connect(lambda _, r=row: self._add_screenshot(r))
        lay.addWidget(btn_add)
        lay.addStretch()
        self.table.setCellWidget(row, COL_IDX["付款截图"], w)

    # ── 合同单元格 ───────────────────────────────
    def _set_contract_cell(self, row, data):
        contracts = data.get("contracts", [])
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(4, 4, 4, 4)
        lay.setSpacing(4)

        if contracts:
            lbl = QLabel(f"[{len(contracts)}]")
            lbl.setStyleSheet(f"color:{GREEN}; font-size:12px; font-weight:bold;")
            btn_v = QPushButton("查看")
            btn_v.setFixedHeight(24)
            btn_v.setFixedWidth(40)
            btn_v.setStyleSheet(f"font-size:11px; padding:1px 4px; color:{GREEN}; border:none; background:transparent;")
            btn_v.clicked.connect(lambda _, r=row: self._view_contracts(r))
            lay.addWidget(lbl)
            lay.addWidget(btn_v)
        else:
            lbl = QLabel("—")
            lbl.setStyleSheet(f"color:{TEXT_DIM}; font-size:12px;")
            lay.addWidget(lbl)

        btn_add = QPushButton("＋")
        btn_add.setFixedHeight(24)
        btn_add.setFixedWidth(26)
        btn_add.setToolTip("添加合同文件（PDF/Word）")
        btn_add.setStyleSheet(f"font-size:13px; padding:0; color:{GREEN}; border:none; background:transparent;")
        btn_add.clicked.connect(lambda _, r=row: self._add_contract(r))
        lay.addWidget(btn_add)
        lay.addStretch()
        self.table.setCellWidget(row, COL_IDX["合同"], w)

    # ── 截图操作 ─────────────────────────────────
    def _get_record_by_row(self, row):
        inv_no_item = self.table.item(row, COL_IDX["发票号码"])
        inv_no = inv_no_item.text() if inv_no_item else ""
        idx = self._find_record_index(inv_no)
        if idx is not None:
            return self.records[idx]
        shown = self._shown_records
        if 0 <= row < len(shown):
            return shown[row]
        return None

    def _add_screenshot(self, row):
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择付款截图", "",
            "图片文件 (*.png *.jpg *.jpeg *.bmp *.gif *.webp);;所有文件 (*)"
        )
        if files:
            self._add_screenshots_from_paths(row, files)

    def _add_screenshots_from_paths(self, row, src_paths):
        rec = self._get_record_by_row(row)
        if rec is None:
            return
        added = self._svc.add_attachments(rec, src_paths, "screenshots",
                                           self._screenshot_dir,
                                           InvoiceService.screenshot_namer)
        if added > 0:
            self._set_screenshot_cell(row, rec)
            self._save_data()
            self.status.showMessage(f"已为该发票添加 {added} 张付款截图")

    def _view_screenshots(self, row):
        rec = self._get_record_by_row(row)
        if rec is None:
            return
        screenshots = rec.get("screenshots", [])
        if not screenshots:
            QMessageBox.information(self, "提示", "该发票暂无付款截图")
            return
        dlg = ImageViewerDialog(screenshots, parent=self)
        dlg.exec_()
        # 同步删除操作
        remaining = dlg.get_remaining_paths()
        if len(remaining) != len(screenshots):
            rec["screenshots"] = remaining
            self._set_screenshot_cell(row, rec)
            self._save_data()

    # ── 合同操作 ─────────────────────────────────
    def _add_contract(self, row):
        """通过文件选择对话框添加合同"""
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择合同文件", "",
            "合同文件 (*.pdf *.docx *.doc *.xlsx *.xls);;所有文件 (*)"
        )
        if files:
            self._add_contracts_from_paths(row, files)

    def _add_contracts_from_paths(self, row, src_paths):
        """将合同文件复制到 contracts 目录并绑定到指定行"""
        rec = self._get_record_by_row(row)
        if rec is None:
            return
        added = self._svc.add_attachments(rec, src_paths, "contracts",
                                           self._contract_dir,
                                           InvoiceService.contract_namer)
        if added > 0:
            self._set_contract_cell(row, rec)
            self._save_data()
            self.status.showMessage(f"已为该发票添加 {added} 份合同")

    # ── 统一附件添加 ─────────────────────────────
    def _add_attachments_from_paths(self, row, src_paths):
        """统一添加附件（图片+文档）"""
        rec = self._get_record_by_row(row)
        if rec is None:
            return
        added = self._svc.add_attachments(rec, src_paths, "attachments",
                                          self._attachment_dir,
                                          InvoiceService._attachment_namer)
        if added > 0:
            # Fallback: use _set_screenshot_cell until _set_attachment_cell is added
            self._set_screenshot_cell(row, rec)
            self._save_data()
            self.status.showMessage(f"已添加 {added} 个附件")

    def _view_contracts(self, row):
        """打开合同管理对话框"""
        rec = self._get_record_by_row(row)
        if rec is None:
            return
        contracts = rec.get("contracts", [])
        if not contracts:
            QMessageBox.information(self, "提示", "该发票暂无合同文件")
            return
        dlg = ContractManagerDialog(
            contracts,
            rec_name=rec.get("buyer_name", "") or rec.get("file", ""),
            parent=self
        )
        dlg.exec_()
        # 同步对话框中可能的删除操作
        if dlg.contract_paths != contracts:
            rec["contracts"] = dlg.contract_paths
            self._set_contract_cell(row, rec)
            self._save_data()

    # ── 剪贴板粘贴 ──────────────────────────────
    def _paste_from_clipboard(self, row):
        """Ctrl+V：图片数据→截图，文件路径→按扩展名分类"""
        clipboard = QApplication.clipboard()
        mime = clipboard.mimeData()

        # 图片像素数据（截图工具）
        if mime.hasImage():
            img = clipboard.image()
            if not img.isNull():
                rec = self._get_record_by_row(row)
                if rec is None:
                    return
                inv_no    = rec.get("invoice_no", "") or rec.get("file", "unnamed")
                safe_name = re.sub(r'[\\/:*?"<>|]', '_', inv_no)
                ts  = datetime.now().strftime("%Y%m%d%H%M%S%f")
                dst = os.path.join(self._attachment_dir, f"{safe_name}_{ts}.png")
                try:
                    img.save(dst, "PNG")
                    rec.setdefault("attachments", []).append(dst)
                    self._set_screenshot_cell(row, rec)
                    self._save_data()
                    self.status.showMessage("已从剪贴板粘贴图片并添加为附件")
                except Exception as ex:
                    QMessageBox.warning(self, "粘贴失败", f"保存剪贴板图片失败：{ex}")
                return

        # 文件路径
        if mime.hasUrls():
            other_files = []
            for u in mime.urls():
                path = u.toLocalFile()
                ext  = os.path.splitext(path)[1].lower()
                if ext in IMG_EXTS or ext in CONTRACT_EXTS:
                    other_files.append(path)
            if other_files:
                self._add_attachments_from_paths(row, other_files)
                return

        self.status.showMessage("剪贴板中没有可用内容（图片或合同文件），请先复制后再粘贴")

    # ── 点击同一行取消选中（viewport 事件过滤器）────
    def eventFilter(self, obj, event):
        if obj is self.table.viewport():
            if event.type() == QEvent.Resize:
                self._recenter_empty_overlay()
            elif event.type() == QEvent.MouseButtonPress:
                if event.button() == Qt.LeftButton:
                    row = self.table.rowAt(event.pos().y())
                    selected = self._selected_rows()
                    if row >= 0 and selected == [row]:
                        self.table.clearSelection()
                        return True
        return super().eventFilter(obj, event)



    # ── 右键菜单 ─────────────────────────────────
    def _show_context_menu(self, pos):
        menu = QMenu(self)

        # 截图区
        menu.addAction(get_icon('camera'), "添加付款截图（文件选择）", self._ctx_add_screenshot)
        menu.addAction(get_icon('clipboard'), "粘贴截图（Ctrl+V）", self._ctx_paste_screenshot)
        menu.addAction(get_icon('search'), "查看付款截图", self._ctx_view_screenshot)
        menu.addAction(get_icon('delete'), "清除此行截图", self._ctx_delete_screenshots)
        menu.addSeparator()

        # 合同区
        menu.addAction(get_icon('document'), "添加合同（文件选择）", self._ctx_add_contract)
        menu.addAction(get_icon('clipboard'), "粘贴合同文件（Ctrl+V）", self._ctx_paste_contract)
        menu.addAction(get_icon('folder'), "查看/管理合同", self._ctx_view_contracts)
        menu.addAction(get_icon('delete'), "清除此行合同", self._ctx_delete_contracts)
        menu.addSeparator()

        menu.addAction(get_icon('delete'), "删除选中行", self._delete_selected_rows)
        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def _selected_rows(self):
        return sorted(set(item.row() for item in self.table.selectedItems()))

    def _ctx_add_screenshot(self):
        for row in self._selected_rows():
            self._add_screenshot(row)

    def _ctx_paste_screenshot(self):
        rows = self._selected_rows()
        if rows:
            self._paste_from_clipboard(rows[0])

    def _ctx_view_screenshot(self):
        rows = self._selected_rows()
        if rows:
            self._view_screenshots(rows[0])

    def _ctx_delete_screenshots(self):
        for row in self._selected_rows():
            rec = self._get_record_by_row(row)
            if rec:
                rec["screenshots"] = []
                self._set_screenshot_cell(row, rec)
        self._save_data()
        self.status.showMessage("已清除选中行的截图记录")

    def _ctx_add_contract(self):
        for row in self._selected_rows():
            self._add_contract(row)

    def _ctx_paste_contract(self):
        rows = self._selected_rows()
        if rows:
            self._paste_from_clipboard(rows[0])

    def _ctx_view_contracts(self):
        rows = self._selected_rows()
        if rows:
            self._view_contracts(rows[0])

    def _ctx_delete_contracts(self):
        for row in self._selected_rows():
            rec = self._get_record_by_row(row)
            if rec:
                rec["contracts"] = []
                self._set_contract_cell(row, rec)
        self._save_data()
        self.status.showMessage("已清除选中行的合同记录")

    def _delete_selected_rows(self):
        rows = sorted(set(item.row() for item in self.table.selectedItems()), reverse=True)
        if not rows:
            return

        # 收集待删除记录信息（先收集再删）
        shown = self._shown_records
        to_delete = []
        for row in rows:
            inv_no_item = self.table.item(row, COL_IDX["发票号码"])
            inv_no = inv_no_item.text() if inv_no_item else ""
            idx = self._find_record_index(inv_no)
            rec = self.records[idx] if idx is not None else None
            if rec is None and 0 <= row < len(shown):
                rec = shown[row]
            if rec and rec not in to_delete:
                to_delete.append(rec)

        if not to_delete:
            return

        # 双重确认弹窗：必须勾选才能点删除
        dlg = DeleteConfirmDialog(to_delete, self)
        if dlg.exec_() != QDialog.Accepted:
            return

        # 删除原始PDF并从 records 移除
        deleted_files = 0
        failed_files  = []
        for rec in to_delete:
            pdf_path = rec.get("pdf_path", "")
            if pdf_path and os.path.exists(pdf_path):
                try:
                    os.remove(pdf_path)
                    deleted_files += 1
                except Exception as ex:
                    failed_files.append(f"{os.path.basename(pdf_path)}：{ex}")
            self.records.remove(rec)

        # 重建表格
        self._rebuild_table()
        self._refresh_filter_combos()
        self._save_data()

        msg = f"已删除 {len(to_delete)} 条记录"
        log.info("删除 %d 条记录 (PDF 成功 %d, 失败 %d)",
                 len(to_delete), deleted_files, len(failed_files))
        if deleted_files:
            msg += f"，{deleted_files} 个PDF文件已删除"
        if failed_files:
            msg += f"，{len(failed_files)} 个文件删除失败"
            QMessageBox.warning(self, "部分文件删除失败",
                "以下文件删除失败（记录已从列表移除）：\n\n" +
                "\n".join(failed_files))
        self.status.showMessage(msg)

    # ── 统计汇总 ─────────────────────────────────
    def _refresh_summary(self):
        self._refresh_summary_from_list(self.records)

    def _refresh_summary_from_list(self, recs):
        count     = len(recs)
        total_amt = sum(safe_float(r.get("amount"))     for r in recs)
        total_tax = sum(safe_float(r.get("tax_amount")) for r in recs)
        total_all = sum(safe_float(r.get("total"))      for r in recs)
        self.lbl_count._value_label.setText(f"{count} 张")
        self.lbl_total_amt._value_label.setText(f"¥ {total_amt:,.2f}")
        self.lbl_total_tax._value_label.setText(f"¥ {total_tax:,.2f}")
        self.lbl_total_all._value_label.setText(f"¥ {total_all:,.2f}")

    def _parse_done(self):
        self._rebuild_table()
        self._refresh_filter_combos()
        self._save_data()

        self.btn_open.setEnabled(True)
        self.progress_bar.setVisible(False)
        batch_count = len(self.records) - getattr(self, '_batch_count_before', 0)
        fail_count = len(getattr(self, '_parse_errors', []))
        dup_count = len(getattr(self, '_duplicate_invoices', []))
        ok_count = max(0, batch_count - fail_count)

        # 结果摘要弹窗（有失败或重复时才弹出）
        if fail_count > 0 or dup_count > 0:
            msg_parts = []
            if ok_count > 0:
                msg_parts.append(f"成功导入 {ok_count} 张")
            if fail_count > 0:
                msg_parts.append(f"解析失败 {fail_count} 张")
            if dup_count > 0:
                msg_parts.append(f"重复跳过 {dup_count} 张")

            detail = "\n\n"
            if fail_count > 0:
                detail += "解析失败：\n" + "\n".join(f"  · {e}" for e in self._parse_errors)
            if dup_count > 0:
                dup_nos = [inv.invoice_no for inv in self._duplicate_invoices]
                detail += "\n重复发票号：\n" + "\n".join(f"  · {n}" for n in dup_nos)

            QMessageBox.information(
                self, "导入结果",
                "\n".join(msg_parts) + detail
            )

        status_msg = " | ".join(msg_parts) if (fail_count > 0 or dup_count > 0) else f"导入完成：{ok_count} 张"
        self.status.showMessage(status_msg)

        # 清理临时状态
        self._parse_errors = []
        self._duplicate_invoices = []

    # ── 导出 Excel ───────────────────────────────
    def export_excel(self):
        export_records = [r for r in self.records if self._record_matches_filter(r)]
        if not export_records:
            QMessageBox.information(self, "提示", "暂无数据，请先导入发票")
            return
        log.info("开始导出 Excel: %d 条记录", len(export_records))

        month_hint = ""
        if self._filter_year or self._filter_month:
            y  = self._filter_year or "全年"
            mo = f"{self._filter_month:02d}月" if self._filter_month else ""
            month_hint = f"_{y}{mo}"

        default_name = f"发票归档{month_hint}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self, "保存Excel文件", default_name, "Excel文件 (*.xlsx)"
        )
        if not save_path:
            return

        # 导出列：不含「付款截图」和「合同」
        xl_columns = [c for c in COLUMNS if c not in ("付款截图", "合同", "操作")]

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "发票归档"

            header_fill  = PatternFill("solid", fgColor="1E6FBF")
            header_font  = Font(color="FFFFFF", bold=True, size=12)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin   = Side(style="thin", color="AAAAAA")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            ws.append(xl_columns)
            for cell in ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = header_align
                cell.border    = border
            ws.row_dimensions[1].height = 28

            alt_fill     = PatternFill("solid", fgColor="EEF4FB")
            normal_align = Alignment(horizontal="left",   vertical="center")
            center_align = Alignment(horizontal="center", vertical="center")

            for i, rec in enumerate(export_records, 2):
                row_data = [
                    rec.get("invoice_type", ""),
                    rec.get("buyer_name", ""),
                    rec.get("buyer_tax_id", ""),
                    rec.get("seller_name", ""),
                    rec.get("amount", ""),
                    rec.get("tax_rate", ""),
                    rec.get("tax_amount", ""),
                    rec.get("total", ""),
                    rec.get("invoice_no", ""),
                    rec.get("invoice_date", ""),
                    rec.get("company", ""),
                    rec.get("remark", "") or rec.get("error", "") or "✓"
                ]
                ws.append(row_data)
                fill = alt_fill if i % 2 == 0 else None
                for j, cell in enumerate(ws[i]):
                    if fill:
                        cell.fill = fill
                    cell.border    = border
                    cell.alignment = center_align if j in [4, 5, 6, 7] else normal_align
                ws.row_dimensions[i].height = 20

            # 汇总行
            ws.append([])
            sum_row   = ws.max_row + 1
            total_amt = sum(safe_float(r.get("amount"))     for r in export_records)
            total_tax = sum(safe_float(r.get("tax_amount")) for r in export_records)
            total_all = sum(safe_float(r.get("total"))      for r in export_records)
            ws.cell(sum_row, 1, "合计")
            ws.cell(sum_row, 5, round(total_amt, 2))
            ws.cell(sum_row, 7, round(total_tax, 2))
            ws.cell(sum_row, 8, round(total_all, 2))
            sum_font = Font(bold=True, color="1E6FBF", size=12)
            sum_fill = PatternFill("solid", fgColor="D6E4F5")
            for cell in ws[sum_row]:
                cell.font      = sum_font
                cell.fill      = sum_fill
                cell.border    = border
                cell.alignment = center_align
            ws.row_dimensions[sum_row].height = 24

            # 列宽：发票类型, 购买方名称, 税号, 销售方名称, 金额, 税率, 税额, 合计, 发票号, 日期, 企业号, 备注
            xl_widths = [16, 20, 22, 20, 12, 8, 12, 14, 20, 14, 15, 14]
            for i, w in enumerate(xl_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            ws.freeze_panes = "A2"
            wb.save(save_path)
            log.info("Excel 导出成功: %s (%d 条)", save_path, len(export_records))
            QMessageBox.information(self, "导出成功",
                f"已成功导出 {len(export_records)} 条记录\n\n路径：{save_path}")
            self.status.showMessage(f"Excel 已保存：{save_path}")
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(save_path)))

        except Exception as e:
            log.error("Excel 导出失败: %s", e, exc_info=True)
            QMessageBox.critical(self, "导出失败", f"导出时出错：\n{e}")

# ─────────────────────────────────────────────
#  入口
# ─────────────────────────────────────────────

def main():
    # 单实例检测
    from PyQt5.QtCore import QSharedMemory
    _singleton = QSharedMemory("lan-invoice-app")
    if not _singleton.create(1) and _singleton.error() == QSharedMemory.AlreadyExists:
        QMessageBox.warning(None, "提示", "发票归档工具已在运行中。\n请查看任务栏或系统托盘。")
        sys.exit(0)

    # 先创建 QApplication（日志 setup 内的 Qt 钩子需要它）
    app = QApplication(sys.argv)
    app.setApplicationName("发票归档")
    app.setStyle("Fusion")
    app.setFont(QFont("Microsoft YaHei", 9))
    app.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    app.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    # 日志初始化
    def _gui_error(title, msg):
        QMessageBox(QMessageBox.Critical, title, msg, QMessageBox.Ok, None).exec_()
    setup_logging(gui_error_callback=_gui_error)

    # 应用级图标（任务栏显示）
    from PyQt5.QtGui import QIcon
    icon_paths = [
        os.path.join(os.path.dirname(__file__), "ui", "icons", "app_icon.png"),
        "icon.ico",
    ]
    for p in icon_paths:
        if os.path.exists(p):
            app.setWindowIcon(QIcon(p))
            break

    win = InvoiceApp()
    win.show()
    exit_code = app.exec_()
    shutdown_logging()
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
