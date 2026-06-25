# -*- coding: utf-8 -*-
"""Excel 导出服务"""

from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models import Invoice
from utils import safe_float
from logger import getLogger

log = getLogger(__name__)

# 固定导出列
_FIXED_COLUMNS = [
    "发票类型", "购买方名称", "纳税人识别号",
    "销售方名称", "金额(元)", "征收率", "税额(元)", "价税合计(元)",
    "发票号码", "开票日期", "备注",
]

_FIELD_MAP = {
    "发票类型": "invoice_type", "购买方名称": "buyer_name",
    "纳税人识别号": "buyer_tax_id", "销售方名称": "seller_name",
    "金额(元)": "amount", "征收率": "tax_rate",
    "税额(元)": "tax_amount", "价税合计(元)": "total",
    "发票号码": "invoice_no", "开票日期": "invoice_date",
}

_BASE_WIDTHS = {
    "发票类型": 16, "购买方名称": 20, "纳税人识别号": 22,
    "销售方名称": 20, "金额(元)": 12, "征收率": 8,
    "税额(元)": 12, "价税合计(元)": 14, "发票号码": 20,
    "开票日期": 14, "备注": 14,
}

_NUMERIC_COLS = {4, 5, 6, 7}  # 金额/征收率/税额/价税合计 列索引（动态生成时重新计算）


class ExportService:
    """将发票列表导出为 Excel 文件"""

    def export(self, invoices: list[Invoice], save_path: str,
               tag_columns: list[str] | None = None) -> None:
        log.info("Excel 导出开始: %d 条 → %s", len(invoices), save_path)
        columns = self._build_columns(tag_columns)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "发票归档"

        self._write_header(ws, columns)
        self._write_data(ws, invoices, columns)
        self._write_summary(ws, invoices, columns)
        self._set_column_widths(ws, columns)

        ws.freeze_panes = "A2"
        wb.save(save_path)
        log.info("Excel 导出完成: %s (%d 条)", save_path, len(invoices))

    @staticmethod
    def _build_columns(tag_columns: list[str] | None) -> list[str]:
        if not tag_columns:
            return list(_FIXED_COLUMNS)
        # 在"备注"前插入标签列
        cols = list(_FIXED_COLUMNS)
        remark_idx = cols.index("备注")
        for tag in reversed(tag_columns):
            cols.insert(remark_idx, tag)
        return cols

    def _write_header(self, ws, columns: list[str]):
        header_fill = PatternFill("solid", fgColor="1E6FBF")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_align = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border
        ws.row_dimensions[1].height = 28

    def _write_data(self, ws, invoices: list[Invoice], columns: list[str]):
        alt_fill = PatternFill("solid", fgColor="EEF4FB")
        normal_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 计算数值列索引
        numeric_indices = set()
        for col_name, col_idx in enumerate(columns):
            if col_name in ("金额(元)", "征收率", "税额(元)", "价税合计(元)"):
                numeric_indices.add(col_idx)

        for i, inv in enumerate(invoices, 2):
            row_data = []
            for col_name in columns:
                if col_name in _FIELD_MAP:
                    row_data.append(getattr(inv, _FIELD_MAP[col_name], ""))
                elif col_name == "备注":
                    row_data.append(inv.remark or inv.error or "✓")
                else:
                    # 标签列
                    row_data.append(inv.tags.get(col_name, ""))
            ws.append(row_data)
            fill = alt_fill if i % 2 == 0 else None
            for j, cell in enumerate(ws[i]):
                if fill:
                    cell.fill = fill
                cell.border = border
                cell.alignment = center_align if j in numeric_indices else normal_align
            ws.row_dimensions[i].height = 20

    def _write_summary(self, ws, invoices: list[Invoice], columns: list[str]):
        ws.append([])
        sum_row = ws.max_row + 1

        total_amt = sum(safe_float(inv.amount) for inv in invoices)
        total_tax = sum(safe_float(inv.tax_amount) for inv in invoices)
        total_all = sum(safe_float(inv.total) for inv in invoices)

        sum_font = Font(bold=True, color="1E6FBF", size=12)
        sum_fill = PatternFill("solid", fgColor="D6E4F5")
        center_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.cell(sum_row, 1, "合计")
        # 金额在"金额(元)"列
        for col_idx, col_name in enumerate(columns, 1):
            if col_name == "金额(元)":
                ws.cell(sum_row, col_idx, round(total_amt, 2))
            elif col_name == "税额(元)":
                ws.cell(sum_row, col_idx, round(total_tax, 2))
            elif col_name == "价税合计(元)":
                ws.cell(sum_row, col_idx, round(total_all, 2))
        for cell in ws[sum_row]:
            cell.font = sum_font
            cell.fill = sum_fill
            cell.border = border
            cell.alignment = center_align
        ws.row_dimensions[sum_row].height = 24

    def _set_column_widths(self, ws, columns: list[str]):
        for i, col_name in enumerate(columns, 1):
            w = _BASE_WIDTHS.get(col_name, 14)  # 标签列默认 14
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
