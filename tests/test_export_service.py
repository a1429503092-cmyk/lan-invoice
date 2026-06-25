# -*- coding: utf-8 -*-
"""export_service 模块单元测试"""
import sys, os, unittest, tempfile, shutil
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

import openpyxl
from models import Invoice
from services.export_service import ExportService


class TestExportService(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.svc = ExportService()

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _make(self, **kw) -> Invoice:
        d = {"file": "test.pdf", "invoice_type": "增值税专用发票",
             "buyer_name": "测试公司", "buyer_tax_id": "12345678901234567",
             "seller_name": "销售方公司", "amount": "550.00", "tax_rate": "13%",
             "tax_amount": "71.50", "total": "621.50",
             "invoice_no": "24113000000012345678",
             "invoice_date": "2024年11月30日", "company": "14786", "remark": ""}
        d.update(kw)
        return Invoice(**d)

    def _export_and_load(self, invoices, tag_columns=None):
        p = os.path.join(self.tmp, "out.xlsx")
        self.svc.export(invoices, p, tag_columns=tag_columns)
        return openpyxl.load_workbook(p).active

    # ── 基本导出 ─────────────────────────────

    def test_export_creates_file(self):
        p = os.path.join(self.tmp, "test.xlsx")
        self.svc.export([self._make()], p)
        self.assertTrue(os.path.exists(p))

    def test_export_single_invoice(self):
        ws = self._export_and_load([self._make(invoice_no="111")])
        self.assertEqual(ws.cell(2, 9).value, "111")  # 发票号码列

    def test_export_multiple_invoices(self):
        invs = [self._make(invoice_no=str(i)) for i in range(1, 6)]
        ws = self._export_and_load(invs)
        self.assertEqual(ws.max_row, 7)  # 5 rows + header + empty + summary

    def test_header_row(self):
        ws = self._export_and_load([self._make()])
        headers = [c.value for c in ws[1]]
        self.assertIn("发票类型", headers)
        self.assertIn("购买方名称", headers)
        self.assertIn("金额(元)", headers)
        self.assertIn("备注", headers)

    # ── 标签列导出 ──────────────────────────

    def test_export_with_tag_columns(self):
        inv = self._make(invoice_no="111", tags={"企业号": "A01", "部门": "研发"})
        ws = self._export_and_load([inv], tag_columns=["企业号", "部门"])
        headers = [c.value for c in ws[1]]
        self.assertIn("企业号", headers)
        self.assertIn("部门", headers)
        # 标签列应在备注前
        tag1_idx = headers.index("企业号")
        tag2_idx = headers.index("部门")
        remark_idx = headers.index("备注")
        self.assertLess(tag1_idx, remark_idx)
        self.assertLess(tag2_idx, remark_idx)
        # 值正确
        self.assertEqual(ws.cell(2, tag1_idx + 1).value, "A01")
        self.assertEqual(ws.cell(2, tag2_idx + 1).value, "研发")

    def test_export_no_tag_columns(self):
        ws = self._export_and_load([self._make()])
        headers = [c.value for c in ws[1]]
        self.assertNotIn("企业号", headers)

    # ── 汇总行 ──────────────────────────────

    def test_summary_row(self):
        invs = [
            self._make(invoice_no="1", amount="100", tax_amount="13", total="113"),
            self._make(invoice_no="2", amount="200", tax_amount="26", total="226"),
        ]
        ws = self._export_and_load(invs)
        last_row = ws.max_row
        self.assertEqual(ws.cell(last_row, 1).value, "合计")
        headers = [c.value for c in ws[1]]
        amt_col = headers.index("金额(元)") + 1
        tax_col = headers.index("税额(元)") + 1
        total_col = headers.index("价税合计(元)") + 1
        self.assertEqual(ws.cell(last_row, amt_col).value, 300)
        self.assertEqual(ws.cell(last_row, tax_col).value, 39)
        self.assertEqual(ws.cell(last_row, total_col).value, 339)

    # ── 红票/错误 ───────────────────────────

    def test_red_invoice_shows_remark(self):
        inv = self._make(invoice_no="1", is_red=True, remark="红票")
        ws = self._export_and_load([inv])
        headers = [c.value for c in ws[1]]
        remark_col = headers.index("备注") + 1
        self.assertEqual(ws.cell(2, remark_col).value, "红票")

    def test_error_shows_in_remark(self):
        inv = self._make(invoice_no="1", error="解析失败", remark="")
        ws = self._export_and_load([inv])
        headers = [c.value for c in ws[1]]
        remark_col = headers.index("备注") + 1
        self.assertEqual(ws.cell(2, remark_col).value, "解析失败")

    def test_freeze_panes(self):
        ws = self._export_and_load([self._make()])
        self.assertEqual(ws.freeze_panes, "A2")


if __name__ == "__main__":
    unittest.main(verbosity=2)
